import markdown
import openpyxl
import pyth.plugins.plaintext.writer
import pyth.plugins.rtf15.reader


# Load the Excel file
workbook = openpyxl.load_workbook("commentaries.xlsx")
sheet = workbook.active

# Iterate through each cell in the sheet
for row in sheet.iter_rows():
    for cell in row:
        # Extract the Markdown content
        try:
            markdown_content = cell.value if cell.value else ""
            html_content = markdown.markdown(markdown_content)
        except Exception as e:
            print(markdown_content)
            import pdb

            pdb.set_trace()
        # Convert Markdown to HTML
        # Convert HTML to RTF
        # doc = pyth.plugins.rtf15.reader.Rtf15Reader.read(html_content)
        # rtf_content = pyth.plugins.plaintext.writer.PlaintextWriter.write(doc).getvalue()
        # Update the cell value with RTF content
        cell.value = html_content

# Save the modified workbook
workbook.save("output.xlsx")


#


import html2text
import openpyxl


# Load the Excel file
workbook = openpyxl.load_workbook("output.xlsx")
sheet = workbook.active

# HTML content
html_content = "<h1>Hello, World!</h1><p>This is <strong>HTML</strong> content.</p>"

# Convert HTML to plain text
plain_text = html2text.html2text(html_content)

# Get the active cell
cell = sheet.active_cell

# Write the plain text content to the cell
cell.value = plain_text

# Apply formatting to the cell
cell.font = openpyxl.styles.Font(bold=True, italic=True, underline="single")

# Save the modified workbook
workbook.save("output2.xlsx")


#######

import openpyxl


# Load the Excel file
workbook = openpyxl.load_workbook("output.xlsx")

# Create a new workbook in write-only mode
output_workbook = openpyxl.Workbook(write_only=True)

# Iterate through each sheet in the input workbook
for sheet_name in workbook.sheetnames:
    # Select the sheet in the input workbook
    sheet = workbook[sheet_name]

    # Create a new sheet in the output workbook
    output_sheet = output_workbook.create_sheet(title=sheet_name)

    # Iterate through each row in the input sheet
    for row in sheet.iter_rows(values_only=True):
        # Write the row to the output sheet without escaping HTML
        output_sheet.append(row)

# Save the output workbook as HTML
output_workbook.save("output.html")
